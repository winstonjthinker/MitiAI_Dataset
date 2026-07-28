import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

processed_dir = r"C:\Users\HP\Downloads\Miti Ai BioBank_v1.0\mitiAI_ZW_Biobank v1.1_Dataset\mitiai_zw_biobank_v1.0_ai4i_data\processed"

tables = [
    ("heritage_knowledge_v1", "Heritage Knowledge (T4)"),
    ("environmental_samples_v1", "Environmental Samples (T1)"),
    ("genomic_sequences_v1", "Genomic Sequences (T2)"),
    ("metabolomic_profiles_v1", "Metabolomic Profiles (T3)"),
    ("qc_log_v1", "Quality Control Log (T5)")
]

header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Navy blue
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
data_font = Font(name="Calibri", size=10)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

def format_worksheet(ws):
    ws.views.sheetView[0].showGridLines = True
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

# 1. Export individual .xlsx for each dataset table
for file_base, title in tables:
    csv_path = os.path.join(processed_dir, f"{file_base}.csv")
    xlsx_path = os.path.join(processed_dir, f"{file_base}.xlsx")
    
    df = pd.read_csv(csv_path)
    df.to_excel(xlsx_path, index=False, engine='openpyxl')
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    ws.title = file_base.replace("_v1", "")
    format_worksheet(ws)
    wb.save(xlsx_path)
    print(f"Saved formatted Excel file: {file_base}.xlsx ({len(df)} rows)")

# 2. Export master all-in-one workbook
master_xlsx = os.path.join(processed_dir, "zw_biobank_v1.0_all_tables.xlsx")
wb_master = openpyxl.Workbook()
# remove default sheet
wb_master.remove(wb_master.active)

for file_base, title in tables:
    csv_path = os.path.join(processed_dir, f"{file_base}.csv")
    df = pd.read_csv(csv_path)
    sheet_name = file_base.replace("_v1", "")
    ws = wb_master.create_sheet(title=sheet_name)
    
    # write headers
    ws.append(list(df.columns))
    # write rows
    for row in df.itertuples(index=False, name=None):
        ws.append([None if pd.isna(x) else x for x in row])
        
    format_worksheet(ws)

wb_master.save(master_xlsx)
print(f"Saved Master Excel Workbook: zw_biobank_v1.0_all_tables.xlsx")
